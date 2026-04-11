#!/usr/bin/env python3
"""
Importa productos desde Excel de inventario del cliente.

Hoja usada (en este orden si existe): **FORMATO INVENTARIO**, si no **Productos** (export generado por la API).

1) **Extendido (Excel del cliente)** — hoja FORMATO INVENTARIO:
   CODIGO, NOMBRE DE PRODUCTO, …, CODIGO FAMILIA, …, PRECIO DE COMPRA, P.V AL PUBLICO, EXISTENCIA.
   La categoría se resuelve con **scripts/familias_cliente.json** (código familia → nombre).

2) **Legado / export app** — categoría en texto; cabeceras típicas:
   Código, Nombre, Categoría, Precio Compra, Precio Venta, Stock Total (hoja **Productos**).

- Crea categorías si no existen (nombre único).
- Códigos duplicados entre filas/archivos: solo cuenta la primera aparición (aviso en stderr).
- Re-ejecutar: UPSERT por código (actualiza nombre, precios, stock, categoría).

Uso:
  pip install psycopg2-binary openpyxl

  python3 scripts/import_productos_excel.py \\
    "/ruta/inventario competo Ke Encanto abril2026.xlsx" \\
    "/ruta/inventario Bonitas 6 abril.xlsx" \\
    "/ruta/INVENTARIO ACTUAL YUSTER.xlsx"

  python3 scripts/import_productos_excel.py -f uno.xlsx -f otro.xlsx

  DATABASE_URL=postgresql://... python3 scripts/import_productos_excel.py --file archivo.xlsx

Sin archivos en línea de comandos, si existen en Descargas los tres Excel del cliente, se usan por defecto.
Para cargar el inventario completo exportado por el sistema (~800+ ítems), pasa también
``inventario_productos_YYYY-MM-DD.xlsx`` (hoja Productos).
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import psycopg2
except ImportError:
    print("Instala: pip install psycopg2-binary openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Instala: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent

DEFAULT_XLSX_CANDIDATES = [
    Path("/home/william/Descargas/inventario competo Ke Encanto abril2026.xlsx"),
    Path("/home/william/Descargas/inventario Bonitas 6 abril.xlsx"),
    Path("/home/william/Descargas/INVENTARIO ACTUAL YUSTER.xlsx"),
]

SHEET_PRIORITY = ("FORMATO INVENTARIO", "Productos")


def pick_sheet_name(wb) -> str | None:
    for name in SHEET_PRIORITY:
        if name in wb.sheetnames:
            return name
    return None


def load_dsn() -> str:
    env = os.environ.get("DATABASE_URL", "").strip()
    if env:
        return env
    cfg = ROOT / "appsettings.json"
    if not cfg.is_file():
        print(f"No hay DATABASE_URL ni {cfg}", file=sys.stderr)
        sys.exit(1)
    with open(cfg, encoding="utf-8") as f:
        data = json.load(f)
    dsn = (data.get("ConnectionStrings") or {}).get("DefaultConnection")
    if not dsn:
        print("ConnectionStrings:DefaultConnection no encontrado.", file=sys.stderr)
        sys.exit(1)
    return dsn.strip()


def load_familias_map() -> dict[int, str]:
    path = SCRIPT_DIR / "familias_cliente.json"
    if not path.is_file():
        print(f"No se encuentra {path}", file=sys.stderr)
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    out: dict[int, str] = {}
    for row in data:
        out[int(row["codigo"])] = str(row["nombre"]).strip()
    return out


def norm_cat(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def to_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    try:
        return Decimal(str(v).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def to_int_stock(v) -> int:
    if v is None or v == "":
        return 0
    try:
        x = float(v)
        return max(0, int(round(x)))
    except (TypeError, ValueError):
        return 0


def truncate(s: str, n: int) -> str:
    s = str(s or "").strip()
    return s[:n] if len(s) > n else s


def header_cells(ws) -> list[str]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(c).strip() if c is not None else "" for c in row1]


def col_index(headers: list[str], *candidates: str) -> int | None:
    norm = [h.strip().upper() for h in headers]
    for cand in candidates:
        c = cand.strip().upper()
        try:
            return norm.index(c)
        except ValueError:
            continue
    return None


def to_familia_codigo(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


SQL_UPSERT_PRODUCT = """
INSERT INTO "Productos" (
  "Codigo", "Nombre", "Descripcion", "Precio", "PrecioCompra",
  "CategoriaProductoId", "StockTotal", "StockMinimo", "ControlarStock",
  "Activo", "FechaCreacion", "FechaActualizacion"
) VALUES (%s, %s, NULL, %s, %s, %s, %s, 0, TRUE, TRUE, NOW(), NOW())
ON CONFLICT ("Codigo") DO UPDATE SET
  "Nombre" = EXCLUDED."Nombre",
  "Precio" = EXCLUDED."Precio",
  "PrecioCompra" = EXCLUDED."PrecioCompra",
  "CategoriaProductoId" = EXCLUDED."CategoriaProductoId",
  "StockTotal" = EXCLUDED."StockTotal",
  "FechaActualizacion" = NOW();
"""


def get_or_create_categoria_id(cur, nombre: str) -> int | None:
    n = norm_cat(nombre)
    if not n:
        return None
    cur.execute(
        'SELECT "Id" FROM "CategoriasProducto" WHERE LOWER(TRIM("Nombre")) = LOWER(%s) LIMIT 1',
        (n,),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])
    cur.execute(
        """
        INSERT INTO "CategoriasProducto" ("Nombre", "Descripcion", "Activo", "FechaCreacion")
        VALUES (%s, NULL, TRUE, NOW())
        RETURNING "Id"
        """,
        (n,),
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def process_sheet(
    ws,
    cur,
    conn,
    familias_map: dict[int, str],
    seen_codes: set[str],
    file_label: str,
    inserted_counter: list[int],
    skipped_dup: list[int],
    skipped_empty: list[int],
    truncated: list[int],
    unknown_fam: set[int],
) -> None:
    headers = header_cells(ws)
    # CODIGO / Código; NOMBRE DE PRODUCTO / Nombre; Categoría con tilde; export app: Precio Venta, Stock Total
    idx_codigo = col_index(headers, "CODIGO", "CÓDIGO")
    idx_nombre = col_index(headers, "NOMBRE DE PRODUCTO", "NOMBRE")
    idx_fam = col_index(headers, "CODIGO FAMILIA")
    idx_cat_text = col_index(headers, "CATEGORIA", "CATEGORÍA")
    idx_pcompra = col_index(headers, "PRECIO DE COMPRA", "PRECIO COMPRA")
    idx_pventa = col_index(headers, "P.V AL PUBLICO", "P.V PUBLICO", "PRECIO VENTA")
    idx_exist = col_index(headers, "EXISTENCIA", "STOCK TOTAL")

    extended = idx_fam is not None and idx_codigo is not None and idx_nombre is not None

    if extended:
        if idx_pcompra is None or idx_pventa is None or idx_exist is None:
            print(
                f"[{file_label}] Formato extendido: faltan columnas "
                f"PRECIO DE COMPRA / P.V AL PUBLICO / EXISTENCIA en cabecera.",
                file=sys.stderr,
            )
            sys.exit(1)
    else:
        legacy_ok = (
            idx_codigo is not None
            and idx_nombre is not None
            and idx_cat_text is not None
            and idx_pcompra is not None
            and idx_pventa is not None
            and idx_exist is not None
        )
        if not legacy_ok:
            print(
                f"[{file_label}] No se reconoce el formato. "
                f"Extendido (CODIGO FAMILIA) o legado/export (Categoría + Precio Compra/Venta + Stock). "
                f"Cabeceras: {headers[:16]}...",
                file=sys.stderr,
            )
            sys.exit(1)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            skipped_empty[0] += 1
            continue
        pad = list(row) + [None] * 40
        codigo_raw = pad[idx_codigo] if idx_codigo is not None else None
        nombre_raw = pad[idx_nombre] if idx_nombre is not None else None

        if codigo_raw is None and nombre_raw is None:
            skipped_empty[0] += 1
            continue

        codigo = truncate(str(codigo_raw).strip(), 50) if codigo_raw is not None else ""
        nombre = truncate(str(nombre_raw).strip(), 200) if nombre_raw is not None else ""

        if not codigo or not nombre:
            skipped_empty[0] += 1
            continue

        if len(str(codigo_raw).strip()) > 50 or len(str(nombre_raw).strip()) > 200:
            truncated[0] += 1

        if codigo in seen_codes:
            skipped_dup[0] += 1
            print(
                f"[dup {file_label} fila {idx}] código repetido (omitido): {codigo}",
                file=sys.stderr,
            )
            continue
        seen_codes.add(codigo)

        if extended:
            fc = to_familia_codigo(pad[idx_fam])
            if fc is None:
                cat_name = "SIN FAMILIA"
            else:
                cat_name = familias_map.get(fc)
                if not cat_name:
                    unknown_fam.add(fc)
                    cat_name = f"FAMILIA {fc}"
        else:
            cat_raw = pad[idx_cat_text]
            cat_name = str(cat_raw) if cat_raw is not None else ""

        cat_id = get_or_create_categoria_id(cur, cat_name)
        precio = to_decimal(pad[idx_pventa])
        pcomp = to_decimal(pad[idx_pcompra])
        stock = to_int_stock(pad[idx_exist])

        cur.execute(
            SQL_UPSERT_PRODUCT,
            (codigo, nombre, precio, pcomp, cat_id, stock),
        )
        inserted_counter[0] += 1

        if inserted_counter[0] % 200 == 0:
            conn.commit()
            print(f"  … {inserted_counter[0]} filas acumuladas")


def main() -> None:
    ap = argparse.ArgumentParser(description="Importar productos desde Excel(s) de inventario.")
    ap.add_argument(
        "files",
        nargs="*",
        type=Path,
        help="Uno o más .xlsx (hoja FORMATO INVENTARIO)",
    )
    ap.add_argument(
        "-f",
        "--file",
        action="append",
        type=Path,
        dest="extra_files",
        help="Otro Excel (puede repetirse; se añade a files)",
    )
    args = ap.parse_args()
    files: list[Path] = list(args.files or [])
    if args.extra_files:
        files.extend(args.extra_files)

    if not files:
        files = [p for p in DEFAULT_XLSX_CANDIDATES if p.is_file()]

    if not files:
        print(
            "Indica al menos un .xlsx:\n"
            "  python3 scripts/import_productos_excel.py archivo1.xlsx archivo2.xlsx\n"
            "O coloca en Descargas los tres inventarios del cliente (nombres por defecto).",
            file=sys.stderr,
        )
        sys.exit(1)

    for p in files:
        if not p.is_file():
            print(f"No existe el archivo: {p}", file=sys.stderr)
            sys.exit(1)

    familias_map = load_familias_map()

    inserted = [0]
    skipped_dup = [0]
    skipped_empty = [0]
    truncated = [0]
    seen_codes: set[str] = set()
    unknown_fam: set[int] = set()

    dsn = load_dsn()
    conn = psycopg2.connect(dsn)
    conn.autocommit = False

    try:
        with conn.cursor() as cur:
            for fp in files:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                try:
                    sn = pick_sheet_name(wb)
                    if not sn:
                        print(
                            f"{fp}: ninguna hoja reconocida {SHEET_PRIORITY}. "
                            f"Hojas: {wb.sheetnames}",
                            file=sys.stderr,
                        )
                        sys.exit(1)
                    ws = wb[sn]
                    print(f"→ {fp.name} [hoja: {sn}] ({fp})")
                    process_sheet(
                        ws,
                        cur,
                        conn,
                        familias_map,
                        seen_codes,
                        fp.name,
                        inserted,
                        skipped_dup,
                        skipped_empty,
                        truncated,
                        unknown_fam,
                    )
                finally:
                    wb.close()
                conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}", file=sys.stderr)
        raise
    finally:
        conn.close()

    print(
        f"Listo: {inserted[0]} productos importados (UPSERT por código). "
        f"Omitidos por código duplicado: {skipped_dup[0]}. "
        f"Filas vacías: {skipped_empty[0]}."
    )
    if truncated[0]:
        print(f"Aviso: {truncated[0]} filas con código o nombre truncados (50 / 200).")
    if unknown_fam:
        print(
            "Aviso: códigos de familia no listados en familias_cliente.json "
            f"(se creó categoría genérica): {sorted(unknown_fam)}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
