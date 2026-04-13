#!/usr/bin/env python3
"""
Importación única de inventario (reemplaza al antiguo seed + import por separado).

1) **Sembrado de categorías** desde ``scripts/familias_cliente.json`` (familias/códigos del Excel extendido).
   Omitir con ``--skip-seed-categorias`` si no aplica.

2) **Productos** desde uno o más .xlsx.

Hoja usada (orden): **FORMATO INVENTARIO**, si no **Productos** (export de la API).

- **Extendido:** CODIGO FAMILIA + ``familias_cliente.json``.
- **Legado / inventario completo / export app:** categoría en texto; columnas tipo
  PRECIO COMPRA, P.V AL PUBLICO / PRECIO VENTA / **p. venta**, EXISTENCIA / STOCK TOTAL.

Sin argumentos, si existe ``inventario completo.xlsx`` en la raíz del proyecto, se importa solo ese archivo.

Uso:
  pip install psycopg2-binary openpyxl

  python3 scripts/clear_database.py --yes
  python3 scripts/importar_inventario.py

  python3 scripts/importar_inventario.py /home/usuario/Descargas/otro_inventario.xlsx

  python3 scripts/importar_inventario.py --skip-seed-categorias solo_productos.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import psycopg2
except ImportError:
    print("Instala: pip install psycopg2-binary openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Instala: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent

# Por defecto: inventario unificado en la raíz del backend (nombre que entregó el cliente)
DEFAULT_XLSX_CANDIDATES = [
    ROOT / "inventario completo.xlsx",
]

SHEET_PRIORITY = ("FORMATO INVENTARIO", "Productos")


def pick_sheet_name(wb) -> str | None:
    for name in SHEET_PRIORITY:
        if name in wb.sheetnames:
            return name
    return None


def load_dsn() -> str:
    env = os.environ.get("DATABASE_URL", "").strip()
    if env:
        return env
    cfg = ROOT / "appsettings.json"
    if not cfg.is_file():
        print(f"No hay DATABASE_URL ni {cfg}", file=sys.stderr)
        sys.exit(1)
    with open(cfg, encoding="utf-8") as f:
        data = json.load(f)
    dsn = (data.get("ConnectionStrings") or {}).get("DefaultConnection")
    if not dsn:
        print("ConnectionStrings:DefaultConnection no encontrado.", file=sys.stderr)
        sys.exit(1)
    return dsn.strip()


def load_familias_map() -> dict[int, str]:
    path = SCRIPT_DIR / "familias_cliente.json"
    if not path.is_file():
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    out: dict[int, str] = {}
    for row in data:
        out[int(row["codigo"])] = str(row["nombre"]).strip()
    return out


def seed_categorias_familias(cur, conn) -> int:
    """Inserta/actualiza categorías desde familias_cliente.json (ON CONFLICT por nombre)."""
    path = SCRIPT_DIR / "familias_cliente.json"
    if not path.is_file():
        print(f"Aviso: no hay {path} — se omiten categorías por familia (solo las que creen los productos).")
        return 0
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    sql = """
    INSERT INTO "CategoriasProducto" ("Nombre", "Descripcion", "Activo", "FechaCreacion")
    VALUES (%s, NULL, TRUE, NOW())
    ON CONFLICT ("Nombre") DO UPDATE SET "Activo" = EXCLUDED."Activo"
    """
    n = 0
    for row in data:
        nombre = str(row["nombre"]).strip()
        if not nombre:
            continue
        cur.execute(sql, (nombre,))
        n += 1
    conn.commit()
    return n


def norm_cat(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def to_decimal(v) -> Decimal:
    """
    Convierte celda Excel a decimal.

    En inventario completo.xlsx, ``p. venta`` mezcla:
    - enteros (910 filas): 300, 850…
    - texto (163 filas): ``C$600.00``, ``C$3,400.00`` (miles con coma),
      ``C$ 1.700,00`` / ``c$592`` (coma decimal estilo NI/EU), comillas, espacios.
    """
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, bool):
        return Decimal("0")
    if isinstance(v, int):
        return Decimal(v)
    if isinstance(v, float):
        return Decimal(str(v))

    s = str(v).strip()
    for _ in range(2):
        s = s.strip().strip('"').strip("'").strip("\u201c\u201d\u2018\u2019")
    s = s.replace("\xa0", " ").replace("\u2009", "").strip()
    if not s:
        return Decimal("0")

    s = re.sub(r"(?i)^\s*c\$\s*", "", s)
    s = s.replace("$", "").strip()
    s = re.sub(r"\s+", "", s)
    if not s:
        return Decimal("0")

    if re.fullmatch(r"-?\d+", s):
        return Decimal(s)

    has_comma = "," in s
    has_dot = "." in s

    if has_comma and has_dot:
        last_c = s.rfind(",")
        last_d = s.rfind(".")
        if last_c > last_d:
            # Europeo / NI: 1.700,00 (punto = miles, coma = decimales)
            s = s.replace(".", "").replace(",", ".")
        else:
            # US: 3,400.00
            s = s.replace(",", "")
    elif has_comma and not has_dot:
        parts = s.split(",")
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) <= 2:
            left = parts[0].replace(".", "")
            s = left + "." + parts[1]
        else:
            s = s.replace(",", "")
    elif has_dot and not has_comma:
        # 600.00 → decimal; 1.700 sin coma → a menudo 1700 córdobas (miles con punto)
        m = re.fullmatch(r"(\d{1,3})\.(\d{3})", s)
        if m:
            s = m.group(1) + m.group(2)
        # si no coincide (ej. 12.50), queda para Decimal

    try:
        d = Decimal(s)
        return d
    except (InvalidOperation, ValueError):
        return Decimal("0")


def to_int_stock(v) -> int:
    if v is None or v == "":
        return 0
    try:
        x = float(v)
        return max(0, int(round(x)))
    except (TypeError, ValueError):
        return 0


def truncate(s: str, n: int) -> str:
    s = str(s or "").strip()
    return s[:n] if len(s) > n else s


def header_cells(ws) -> list[str]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(c).strip() if c is not None else "" for c in row1]


def col_index(headers: list[str], *candidates: str) -> int | None:
    norm = [h.strip().upper() for h in headers]
    for cand in candidates:
        c = cand.strip().upper()
        try:
            return norm.index(c)
        except ValueError:
            continue
    return None


def to_familia_codigo(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


SQL_UPSERT_PRODUCT = """
INSERT INTO "Productos" (
  "Codigo", "Nombre", "Descripcion", "Precio", "PrecioCompra",
  "CategoriaProductoId", "StockTotal", "StockMinimo", "ControlarStock",
  "Activo", "FechaCreacion", "FechaActualizacion"
) VALUES (%s, %s, NULL, %s, %s, %s, %s, 0, TRUE, TRUE, NOW(), NOW())
ON CONFLICT ("Codigo") DO UPDATE SET
  "Nombre" = EXCLUDED."Nombre",
  "Precio" = EXCLUDED."Precio",
  "PrecioCompra" = EXCLUDED."PrecioCompra",
  "CategoriaProductoId" = EXCLUDED."CategoriaProductoId",
  "StockTotal" = EXCLUDED."StockTotal",
  "FechaActualizacion" = NOW();
"""


def get_or_create_categoria_id(cur, nombre: str) -> int | None:
    n = norm_cat(nombre)
    if not n:
        return None
    cur.execute(
        'SELECT "Id" FROM "CategoriasProducto" WHERE LOWER(TRIM("Nombre")) = LOWER(%s) LIMIT 1',
        (n,),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])
    cur.execute(
        """
        INSERT INTO "CategoriasProducto" ("Nombre", "Descripcion", "Activo", "FechaCreacion")
        VALUES (%s, NULL, TRUE, NOW())
        RETURNING "Id"
        """,
        (n,),
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def process_sheet(
    ws,
    cur,
    conn,
    familias_map: dict[int, str],
    seen_codes: set[str],
    file_label: str,
    inserted_counter: list[int],
    skipped_dup: list[int],
    skipped_empty: list[int],
    truncated: list[int],
    unknown_fam: set[int],
) -> None:
    headers = header_cells(ws)
    # CODIGO / Código; NOMBRE DE PRODUCTO / Nombre; Categoría con tilde; export app: Precio Venta, Stock Total
    idx_codigo = col_index(headers, "CODIGO", "CÓDIGO")
    idx_nombre = col_index(headers, "NOMBRE DE PRODUCTO", "NOMBRE")
    idx_fam = col_index(headers, "CODIGO FAMILIA")
    idx_cat_text = col_index(headers, "CATEGORIA", "CATEGORÍA")
    idx_pcompra = col_index(headers, "PRECIO DE COMPRA", "PRECIO COMPRA")
    idx_pventa = col_index(headers, "P.V AL PUBLICO", "P.V PUBLICO", "PRECIO VENTA", "P. VENTA")
    idx_exist = col_index(headers, "EXISTENCIA", "STOCK TOTAL")

    extended = idx_fam is not None and idx_codigo is not None and idx_nombre is not None

    if extended:
        if idx_pcompra is None or idx_pventa is None or idx_exist is None:
            print(
                f"[{file_label}] Formato extendido: faltan columnas "
                f"PRECIO DE COMPRA / P.V AL PUBLICO / EXISTENCIA en cabecera.",
                file=sys.stderr,
            )
            sys.exit(1)
    else:
        legacy_ok = (
            idx_codigo is not None
            and idx_nombre is not None
            and idx_cat_text is not None
            and idx_pcompra is not None
            and idx_pventa is not None
            and idx_exist is not None
        )
        if not legacy_ok:
            print(
                f"[{file_label}] No se reconoce el formato. "
                f"Extendido (CODIGO FAMILIA) o legado/export (Categoría + Precio Compra/Venta + Stock). "
                f"Cabeceras: {headers[:16]}...",
                file=sys.stderr,
            )
            sys.exit(1)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            skipped_empty[0] += 1
            continue
        pad = list(row) + [None] * 40
        codigo_raw = pad[idx_codigo] if idx_codigo is not None else None
        nombre_raw = pad[idx_nombre] if idx_nombre is not None else None

        if codigo_raw is None and nombre_raw is None:
            skipped_empty[0] += 1
            continue

        codigo = truncate(str(codigo_raw).strip(), 50) if codigo_raw is not None else ""
        nombre = truncate(str(nombre_raw).strip(), 200) if nombre_raw is not None else ""

        if not codigo or not nombre:
            skipped_empty[0] += 1
            continue

        if len(str(codigo_raw).strip()) > 50 or len(str(nombre_raw).strip()) > 200:
            truncated[0] += 1

        if codigo in seen_codes:
            skipped_dup[0] += 1
            print(
                f"[dup {file_label} fila {idx}] código repetido (omitido): {codigo}",
                file=sys.stderr,
            )
            continue
        seen_codes.add(codigo)

        if extended:
            fc = to_familia_codigo(pad[idx_fam])
            if fc is None:
                cat_name = "SIN FAMILIA"
            else:
                cat_name = familias_map.get(fc)
                if not cat_name:
                    unknown_fam.add(fc)
                    cat_name = f"FAMILIA {fc}"
        else:
            cat_raw = pad[idx_cat_text]
            cat_name = str(cat_raw) if cat_raw is not None else ""

        cat_id = get_or_create_categoria_id(cur, cat_name)
        precio = to_decimal(pad[idx_pventa])
        pcomp = to_decimal(pad[idx_pcompra])
        stock = to_int_stock(pad[idx_exist])

        cur.execute(
            SQL_UPSERT_PRODUCT,
            (codigo, nombre, precio, pcomp, cat_id, stock),
        )
        inserted_counter[0] += 1

        if inserted_counter[0] % 200 == 0:
            conn.commit()
            print(f"  … {inserted_counter[0]} filas acumuladas")


def main() -> None:
    ap = argparse.ArgumentParser(description="Sembrar categorías (JSON) e importar productos desde Excel.")
    ap.add_argument(
        "files",
        nargs="*",
        type=Path,
        help="Uno o más .xlsx",
    )
    ap.add_argument(
        "-f",
        "--file",
        action="append",
        type=Path,
        dest="extra_files",
        help="Otro Excel (puede repetirse)",
    )
    ap.add_argument(
        "--skip-seed-categorias",
        action="store_true",
        help="No ejecutar INSERT de categorías desde familias_cliente.json",
    )
    args = ap.parse_args()
    files: list[Path] = list(args.files or [])
    if args.extra_files:
        files.extend(args.extra_files)

    if not files:
        files = [p for p in DEFAULT_XLSX_CANDIDATES if p.is_file()]

    if not files:
        print(
            "Indica al menos un .xlsx o coloca ``inventario completo.xlsx`` en la raíz del proyecto.\n"
            "  python3 scripts/importar_inventario.py archivo.xlsx",
            file=sys.stderr,
        )
        sys.exit(1)

    for p in files:
        if not p.is_file():
            print(
                f"No existe el archivo: {p}\n"
                "  Usa una ruta real (no el ejemplo /ruta/a/tu.xlsx de la ayuda).\n"
                "  Sin argumentos se usa «inventario completo.xlsx» en la raíz del proyecto si existe.",
                file=sys.stderr,
            )
            sys.exit(1)

    familias_map = load_familias_map()

    inserted = [0]
    skipped_dup = [0]
    skipped_empty = [0]
    truncated = [0]
    seen_codes: set[str] = set()
    unknown_fam: set[int] = set()

    dsn = load_dsn()
    conn = psycopg2.connect(dsn)
    conn.autocommit = False

    try:
        with conn.cursor() as cur:
            if not args.skip_seed_categorias:
                ncat = seed_categorias_familias(cur, conn)
                if ncat:
                    print(f"Categorías desde familias_cliente.json: {ncat} (nuevas o actualizadas).")
            for fp in files:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                try:
                    sn = pick_sheet_name(wb)
                    if not sn:
                        print(
                            f"{fp}: ninguna hoja reconocida {SHEET_PRIORITY}. "
                            f"Hojas: {wb.sheetnames}",
                            file=sys.stderr,
                        )
                        sys.exit(1)
                    ws = wb[sn]
                    print(f"→ {fp.name} [hoja: {sn}] ({fp})")
                    process_sheet(
                        ws,
                        cur,
                        conn,
                        familias_map,
                        seen_codes,
                        fp.name,
                        inserted,
                        skipped_dup,
                        skipped_empty,
                        truncated,
                        unknown_fam,
                    )
                finally:
                    wb.close()
                conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}", file=sys.stderr)
        raise
    finally:
        conn.close()

    print(
        f"Listo: {inserted[0]} productos importados (UPSERT por código). "
        f"Omitidos por código duplicado: {skipped_dup[0]}. "
        f"Filas vacías: {skipped_empty[0]}."
    )
    if truncated[0]:
        print(f"Aviso: {truncated[0]} filas con código o nombre truncados (50 / 200).")
    if unknown_fam:
        print(
            "Aviso: códigos de familia no listados en familias_cliente.json "
            f"(se creó categoría genérica): {sorted(unknown_fam)}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
